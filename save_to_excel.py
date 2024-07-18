import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill

def save_to_excel(df, file_path, length_threshold=20, exclude_columns=[]):
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        ws = writer.sheets['Sheet1']
        
        # Estilizar o cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar a largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Ajustar a altura das linhas, exceto o cabeçalho
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            ws.row_dimensions[row[0].row].height = 50
        
        # Ajustar a largura das colunas com conteúdo excedente, exceto colunas excluídas
        for col in ws.columns:
            col_name = ws.cell(row=1, column=col[0].column).value
            if col_name in exclude_columns:
                continue
            max_length = max(len(str(cell.value)) for cell in col if cell.value)
            if max_length > length_threshold:
                column = col[0].column_letter
                ws.column_dimensions[column].width = max_length + 2

        # Ajustar a coluna 'Descrição' para ter texto empilhado e justificado
        descricao_col_idx = df.columns.get_loc('Descrição') + 1
        descricao_col_letter = ws.cell(row=1, column=descricao_col_idx).column_letter
        ws.column_dimensions[descricao_col_letter].width = 30  # Ajustar a largura conforme necessário
        
        for cell in ws[descricao_col_letter]:
            cell.alignment = Alignment(horizontal='justify', vertical='top', wrap_text=True)


